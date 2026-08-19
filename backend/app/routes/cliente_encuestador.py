import json
import io
import hashlib
import logging
import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import or_, and_, desc, func, case, text, over as sql_over, select
from typing import List, Any, Optional
from datetime import date, datetime, timedelta

from app.db.session import get_db, get_async_db
from app.core.dependencies import get_current_user
from app.models.user import Usuario as User
from app.models.encuestador import JornadaEncuestador, CentroSalud, EncuestaCentro, Medico, MedicoCentroEncuesta, MedicoConsultorio
from app.services.cache_service import cache_get, cache_set

logger = logging.getLogger("app")
router = APIRouter(prefix="/api/cliente-encuestador", tags=["Cliente Encuestador"])

# Mismas keys abreviadas que medico-form.component.ts usa en el JSON de
# horarios (diasList) -- el filtro/gráfico de días tiene que buscar estas
# mismas keys en horarios_json o nunca va a matchear nada.
DIAS_ABREV = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]

async def check_rol_cliente_encuestador(current_user: User, db: AsyncSession):
    # id_rol=13 (IQVIA) y admin siempre pasan. Además, cualquier usuario con
    # el permiso 'cliente-encuestador' concedido a mano (Permisos -> por
    # usuario) también entra, sin importar su rol -- ej. un supervisor del
    # equipo de encuestadores (rol 12) al que se le da acceso de lectura al
    # BI sin volverlo IQVIA ni crear un rol nuevo solo para eso.
    # No se usa current_user.has_permission()/.permisos: esa relación es
    # lazy="noload" y get_current_user no la carga -- siempre saldría vacía.
    # Se consulta usuario_permisos directo, mismo patrón que require_permission()
    # en core/dependencies.py (el único camino que de verdad funciona hoy).
    if current_user.id_rol == 13 or current_user.is_admin:
        return
    perm = (await db.execute(text("""
        SELECT can_read FROM usuario_permisos WHERE id_usuario = :uid AND module = 'cliente-encuestador'
    """), {"uid": current_user.id})).fetchone()
    if perm and perm[0]:
        return
    raise HTTPException(status_code=403, detail="Acceso denegado. Solo para Cliente Encuestador.")

def _dias_activos_str(horarios_json: Optional[str]) -> str:
    if not horarios_json:
        return ''
    try:
        h = json.loads(horarios_json)
    except (TypeError, ValueError):
        return ''
    return ', '.join(d for d in DIAS_ABREV if isinstance(h, dict) and h.get(d, {}).get('activo'))

def _primer_consultorio_alias(db=None):
    min_ids = (
        select(
            MedicoConsultorio.id_medico.label('id_medico'),
            func.min(MedicoConsultorio.id_consultorio).label('min_id')
        )
        .group_by(MedicoConsultorio.id_medico)
        .subquery()
    )
    return (
        select(MedicoConsultorio)
        .join(min_ids, MedicoConsultorio.id_consultorio == min_ids.c.min_id)
        .subquery()
    )


def _latest_medico_centro_subq(db=None):
    rn = (
        select(
            MedicoCentroEncuesta.id_medico_centro,
            func.row_number().over(
                partition_by=Medico.id_medico,
                order_by=[desc(EncuestaCentro.fecha_verificacion), desc(MedicoCentroEncuesta.id_medico_centro)]
            ).label('rn')
        )
        .join(EncuestaCentro, EncuestaCentro.id_encuesta == MedicoCentroEncuesta.id_encuesta)
        .join(Medico, Medico.id_medico == MedicoCentroEncuesta.id_medico)
        .subquery()
    )
    return select(rn.c.id_medico_centro).filter(rn.c.rn == 1).subquery()


def get_base_query(db=None):
    pc = _primer_consultorio_alias(db)
    stmt = (
        select(MedicoCentroEncuesta, EncuestaCentro, Medico, CentroSalud, User, JornadaEncuestador)
        .join(EncuestaCentro, EncuestaCentro.id_encuesta == MedicoCentroEncuesta.id_encuesta)
        .join(Medico, Medico.id_medico == MedicoCentroEncuesta.id_medico)
        .join(CentroSalud, CentroSalud.id_centro == EncuestaCentro.id_centro)
        .join(User, User.id == EncuestaCentro.id_usuario)
        .outerjoin(JornadaEncuestador, JornadaEncuestador.id_jornada == EncuestaCentro.id_jornada)
        .outerjoin(pc, pc.c.id_medico == Medico.id_medico)
    )
    return stmt, pc

def apply_filters(query, req: Request, pc):
    q_params = req.query_params
    
    fdesde = q_params.get("fecha_desde")
    fhasta = q_params.get("fecha_hasta")
    if fdesde: query = query.filter(EncuestaCentro.fecha_verificacion >= fdesde)
    if fhasta: query = query.filter(EncuestaCentro.fecha_verificacion <= fhasta)
    
    def apply_in(col, param_name):
        nonlocal query
        vals = q_params.getlist(param_name)
        if len(vals) == 1 and ',' in vals[0]:
            vals = [v.strip() for v in vals[0].split(',')]
        vals = [v for v in vals if v]
        if vals:
            if param_name in ("centros", "encuestadores"):
                int_vals = []
                for v in vals:
                    try:
                        int_vals.append(int(v))
                    except (ValueError, TypeError):
                        pass
                if int_vals:
                    query = query.filter(col.in_(int_vals))
            else:
                query = query.filter(col.in_(vals))
        return query

    apply_in(Medico.estado, "estados")
    apply_in(Medico.ciudad, "ciudades")
    apply_in(Medico.especialidad, "especialidades")
    apply_in(Medico.sub_especialidad, "sub_especialidades")
    apply_in(Medico.universidad_graduacion, "universidades")
    apply_in(CentroSalud.id_centro, "centros")
    apply_in(EncuestaCentro.id_usuario, "encuestadores")
    apply_in(EncuestaCentro.fuente_informacion, "fuentes")
    apply_in(pc.c.valor_consulta_rango, "valor_consulta_rangos")
    apply_in(pc.c.promedio_pacientes_semanal_rango, "promedio_pacientes_rangos")

    dias = q_params.getlist("dias_consulta")
    if len(dias) == 1 and ',' in dias[0]: dias = [d.strip() for d in dias[0].split(',')]
    dias = [d for d in dias if d]
    if dias:
        ors = [pc.c.horarios_json.ilike(f'%"{d}":{{"activo":true%') for d in dias]
        query = query.filter(or_(*ors))

    return query

@router.get("/filtros")
async def api_filtros(db: AsyncSession = Depends(get_async_db), current_user: User = Depends(get_current_user)):
    await check_rol_cliente_encuestador(current_user, db)
    
    especialidades = [r for r in (await db.execute(select(Medico.especialidad).distinct().filter(Medico.especialidad != None).order_by(Medico.especialidad))).scalars().all() if r]
    sub_especialidades = [r for r in (await db.execute(select(Medico.sub_especialidad).distinct().filter(Medico.sub_especialidad != None).order_by(Medico.sub_especialidad))).scalars().all() if r]
    estados = [r for r in (await db.execute(select(Medico.estado).distinct().filter(Medico.estado != None).order_by(Medico.estado))).scalars().all() if r]
    ciudades = [r for r in (await db.execute(select(Medico.ciudad).distinct().filter(Medico.ciudad != None).order_by(Medico.ciudad))).scalars().all() if r]
    universidades = [r for r in (await db.execute(select(Medico.universidad_graduacion).distinct().filter(Medico.universidad_graduacion != None).order_by(Medico.universidad_graduacion))).scalars().all() if r]
    
    centros = [{"id_centro": r.id_centro, "nombre_centro": r.nombre_centro} for r in (await db.execute(select(CentroSalud.id_centro, CentroSalud.nombre_centro).order_by(CentroSalud.nombre_centro))).fetchall()]
    encuestadores = [{"id_usuario": r.id, "username": r.username} for r in (await db.execute(select(User.id, User.username).join(EncuestaCentro, EncuestaCentro.id_usuario == User.id).distinct().order_by(User.username))).fetchall()]
    
    fuentes = [r for r in (await db.execute(select(EncuestaCentro.fuente_informacion).distinct().filter(EncuestaCentro.fuente_informacion != None).order_by(EncuestaCentro.fuente_informacion))).scalars().all() if r]
    valor_rangos = [r for r in (await db.execute(select(MedicoConsultorio.valor_consulta_rango).distinct().filter(MedicoConsultorio.valor_consulta_rango != None).order_by(MedicoConsultorio.valor_consulta_rango))).scalars().all() if r]
    pac_rangos = [r for r in (await db.execute(select(MedicoConsultorio.promedio_pacientes_semanal_rango).distinct().filter(MedicoConsultorio.promedio_pacientes_semanal_rango != None).order_by(MedicoConsultorio.promedio_pacientes_semanal_rango))).scalars().all() if r]

    return {
        "success": True,
        "especialidades": especialidades,
        "sub_especialidades": sub_especialidades,
        "estados": estados,
        "ciudades": ciudades,
        "universidades": universidades,
        "centros": centros,
        "encuestadores": encuestadores,
        "fuentes": fuentes,
        "valor_consulta_rangos": valor_rangos,
        "promedio_pacientes_rangos": pac_rangos,
        "dias_consulta": DIAS_ABREV
    }

@router.get("/kpis")
async def api_kpis(request: Request, db: AsyncSession = Depends(get_async_db), current_user: User = Depends(get_current_user)):
    await check_rol_cliente_encuestador(current_user, db)

    # 1. Comprobar Caché Redis (Fallback inmediato en <1ms)
    params_str = str(sorted(request.query_params.multi_items()))
    cache_key = f"cliente_encuestador_kpis:{hashlib.md5(params_str.encode()).hexdigest()}"
    cached = await cache_get(cache_key)
    if cached:
        return cached

    try:
        q, pc = get_base_query(db)
        q = apply_filters(q, request, pc)

        # Batching de conteos generales
        totals_row = (await db.execute(q.with_only_columns(
            func.count(func.distinct(Medico.id_medico)),
            func.count(func.distinct(CentroSalud.id_centro)),
            func.count(func.distinct(Medico.especialidad)),
            func.count(func.distinct(Medico.estado)),
            func.count(func.distinct(Medico.ciudad)),
            func.count(func.distinct(EncuestaCentro.id_encuesta))
        ))).fetchone()

        total_medicos = totals_row[0] if totals_row else 0
        total_centros = totals_row[1] if totals_row else 0
        total_especialidades = totals_row[2] if totals_row else 0
        total_estados = totals_row[3] if totals_row else 0
        total_ciudades = totals_row[4] if totals_row else 0
        total_encuestas = totals_row[5] if totals_row else 0

        thirty_days_ago = datetime.utcnow().date() - timedelta(days=30)
        encuestas_30d = (await db.execute(q.filter(EncuestaCentro.fecha_verificacion >= thirty_days_ago).with_only_columns(func.count(func.distinct(EncuestaCentro.id_encuesta))))).scalar() or 0

        # Conteo de 2do consultorio
        medico_ids = [r[0] for r in (await db.execute(q.with_only_columns(Medico.id_medico).distinct())).all() if r[0] is not None]
        dos_cons = 0
        if medico_ids:
            sub = (
                select(MedicoConsultorio.id_medico)
                .filter(MedicoConsultorio.id_medico.in_(medico_ids))
                .group_by(MedicoConsultorio.id_medico)
                .having(func.count(MedicoConsultorio.id_consultorio) > 1)
                .subquery()
            )
            dos_cons = (await db.execute(select(func.count()).select_from(sub))).scalar() or 0
        pct_dos = round((dos_cons * 100.0) / total_medicos, 1) if total_medicos else 0.0

        # Batching de canales de contacto en una sola consulta
        contacts_row = (await db.execute(q.with_only_columns(
            func.count(func.distinct(case((and_(Medico.whatsapp != None, Medico.whatsapp != ''), Medico.id_medico), else_=None))),
            func.count(func.distinct(case((and_(Medico.email != None, Medico.email != ''), Medico.id_medico), else_=None))),
            func.count(func.distinct(case((and_(Medico.telefono != None, Medico.telefono != ''), Medico.id_medico), else_=None))),
            func.count(func.distinct(case((and_(Medico.instagram != None, Medico.instagram != ''), Medico.id_medico), else_=None))),
            func.count(func.distinct(case((and_(Medico.linkedin != None, Medico.linkedin != ''), Medico.id_medico), else_=None)))
        ))).fetchone()

        wa = contacts_row[0] if contacts_row else 0
        em = contacts_row[1] if contacts_row else 0
        tel = contacts_row[2] if contacts_row else 0
        ig = contacts_row[3] if contacts_row else 0
        li = contacts_row[4] if contacts_row else 0

        def pct(x): return round((x * 100.0) / total_medicos, 1) if total_medicos else 0.0

        # --- CHART DATA ---
        esp_data = (await db.execute(q.with_only_columns(Medico.especialidad, func.count(func.distinct(Medico.id_medico))).group_by(Medico.especialidad))).fetchall()
        esp_chart = sorted([{"name": r[0] or "N/A", "value": r[1]} for r in esp_data], key=lambda x: -x["value"])

        est_data = (await db.execute(q.with_only_columns(Medico.estado, func.count(func.distinct(Medico.id_medico))).group_by(Medico.estado))).fetchall()
        est_chart = sorted([{"name": r[0] or "N/A", "value": r[1]} for r in est_data], key=lambda x: -x["value"])

        uni_data = (await db.execute(q.with_only_columns(Medico.universidad_graduacion, func.count(func.distinct(Medico.id_medico))).group_by(Medico.universidad_graduacion))).fetchall()
        uni_chart = sorted([{"name": r[0] or "N/A", "value": r[1]} for r in uni_data], key=lambda x: -x["value"])

        cen_data = (await db.execute(q.with_only_columns(CentroSalud.nombre_centro, func.count(func.distinct(Medico.id_medico))).group_by(CentroSalud.nombre_centro))).fetchall()
        cen_chart = sorted([{"name": r[0] or "N/A", "value": r[1]} for r in cen_data], key=lambda x: -x["value"])

        val_data = (await db.execute(q.with_only_columns(pc.c.valor_consulta_rango, func.count(func.distinct(Medico.id_medico))).group_by(pc.c.valor_consulta_rango))).fetchall()
        val_chart = [{"name": r[0] or "N/A", "value": r[1]} for r in val_data]

        pac_data = (await db.execute(q.with_only_columns(pc.c.promedio_pacientes_semanal_rango, func.count(func.distinct(Medico.id_medico))).group_by(pc.c.promedio_pacientes_semanal_rango))).fetchall()
        pac_chart = [{"name": r[0] or "N/A", "value": r[1]} for r in pac_data]

        enc_data = (await db.execute(q.with_only_columns(User.username, func.count(func.distinct(Medico.id_medico)), func.count(func.distinct(CentroSalud.id_centro)), func.count(func.distinct(EncuestaCentro.id_encuesta))).group_by(User.username))).fetchall()
        enc_ranking = [{"encuestador": r[0], "medicos": r[1], "centros": r[2], "encuestas": r[3]} for r in enc_data]

        dias_data = (await db.execute(q.with_only_columns(pc.c.horarios_json))).fetchall()
        dias_count = {d: 0 for d in DIAS_ABREV}
        horas_count = {h: 0 for h in range(24)}
        for (horarios_str,) in dias_data:
            if not horarios_str:
                continue
            try:
                h = json.loads(horarios_str)
            except (TypeError, ValueError):
                continue
            if not isinstance(h, dict):
                continue
            for d in DIAS_ABREV:
                info = h.get(d) or {}
                if not info.get('activo'):
                    continue
                dias_count[d] += 1
                try:
                    desde_h = int(str(info.get('desde', '00:00')).split(':')[0])
                    hasta_h = int(str(info.get('hasta', '00:00')).split(':')[0])
                except (ValueError, IndexError):
                    continue
                for hh in range(desde_h, min(hasta_h, 24)):
                    if 0 <= hh < 24:
                        horas_count[hh] += 1
        dias_chart = [{"name": k, "value": v} for k, v in dias_count.items()]
        horas_chart = [{"name": f"{h:02d}:00", "value": v} for h, v in horas_count.items()]

        response_data = {
            "success": True,
            "total_medicos": total_medicos,
            "total_centros": total_centros,
            "total_especialidades": total_especialidades,
            "total_estados": total_estados,
            "total_ciudades": total_ciudades,
            "total_encuestas": total_encuestas,
            "encuestas_30d": encuestas_30d,
            "medicos_con_2do_consultorio": dos_cons,
            "pct_2do_consultorio": pct_dos,
            "pct_whatsapp": pct(wa),
            "pct_email": pct(em),
            "pct_telefono": pct(tel),
            "pct_instagram": pct(ig),
            "pct_linkedin": pct(li),
            "charts": {
                "especialidades": esp_chart,
                "estados": est_chart,
                "universidades": uni_chart,
                "centros": cen_chart,
                "valor_consulta": val_chart,
                "pacientes_semana": pac_chart,
                "dias_consulta": dias_chart,
                "horas_consulta": horas_chart,
                "ranking_encuestadores": enc_ranking
            }
        }
        await cache_set(cache_key, response_data, ttl_seconds=120)
        return response_data
    except Exception as e:
        logger.error(f"[cliente_encuestador] Error calculando KPIs (fallback seguro activado): {e}")
        return {
            "success": True,
            "total_medicos": 0, "total_centros": 0, "total_especialidades": 0,
            "total_estados": 0, "total_ciudades": 0, "total_encuestas": 0,
            "encuestas_30d": 0, "medicos_con_2do_consultorio": 0, "pct_2do_consultorio": 0.0,
            "pct_whatsapp": 0.0, "pct_email": 0.0, "pct_telefono": 0.0, "pct_instagram": 0.0, "pct_linkedin": 0.0,
            "charts": {
                "especialidades": [], "estados": [], "universidades": [],
                "centros": [], "valor_consulta": [], "pacientes_semana": [],
                "dias_consulta": [{"name": k, "value": 0} for k in DIAS_ABREV],
                "horas_consulta": [{"name": f"{h:02d}:00", "value": 0} for h in range(24)],
                "ranking_encuestadores": []
            }
        }

@router.get("/medicos")
async def api_medicos_tabla(request: Request, q: str = "", page: int = 1, per_page: int = 25, db: AsyncSession = Depends(get_async_db), current_user: User = Depends(get_current_user)):
    await check_rol_cliente_encuestador(current_user, db)

    try:
        base_q, pc = get_base_query(db)
        base_q = apply_filters(base_q, request, pc)

        # Deduplicar: solo la encuesta más reciente por médico
        dedup = _latest_medico_centro_subq(db)
        base_q = base_q.filter(MedicoCentroEncuesta.id_medico_centro.in_(dedup))

        if q.strip():
            search = f"%{q.strip()}%"
            base_q = base_q.filter(
                or_(
                    Medico.id_medico_externo.ilike(search),
                    Medico.apellido1.ilike(search),
                    Medico.apellido2.ilike(search),
                    Medico.nombre1.ilike(search),
                    Medico.nombre2.ilike(search),
                    Medico.especialidad.ilike(search),
                    CentroSalud.nombre_centro.ilike(search)
                )
            )
            
        total = (await db.execute(base_q.with_only_columns(func.count(func.distinct(Medico.id_medico))))).scalar() or 0
        offset = (page - 1) * per_page
        
        rows = (await db.execute(base_q.with_only_columns(
            Medico.id_medico, Medico.id_medico_externo,
            Medico.apellido1, Medico.apellido2, Medico.nombre1, Medico.nombre2,
            Medico.especialidad, Medico.sub_especialidad, Medico.universidad_graduacion,
            Medico.ciudad, Medico.estado, Medico.telefono, Medico.whatsapp, Medico.email,
            CentroSalud.nombre_centro, pc.c.valor_consulta_rango,
            pc.c.promedio_pacientes_semanal_rango, pc.c.horarios_json,
            EncuestaCentro.fecha_verificacion, User.username,
            JornadaEncuestador.latitud, JornadaEncuestador.longitud
        ).order_by(desc(EncuestaCentro.fecha_verificacion), Medico.apellido1).offset(offset).limit(per_page))).fetchall()

        medicos = []
        for r in rows:
            n2 = f" {r.nombre2}" if r.nombre2 else ""
            a2 = f" {r.apellido2}" if r.apellido2 else ""
            nombre_completo = f"{r.apellido1}{a2}, {r.nombre1}{n2}"
            medicos.append({
                "id_medico": r.id_medico,
                "id_medico_externo": r.id_medico_externo,
                "nombre_completo": nombre_completo,
                "especialidad": r.especialidad,
                "sub_especialidad": r.sub_especialidad,
                "universidad": r.universidad_graduacion,
                "ciudad": r.ciudad,
                "estado": r.estado,
                "telefono": r.telefono,
                "whatsapp": r.whatsapp,
                "email": r.email,
                "centro": r.nombre_centro,
                "valor_consulta_rango": r.valor_consulta_rango,
                "promedio_pacientes": r.promedio_pacientes_semanal_rango,
                "dias_consulta": _dias_activos_str(r.horarios_json),
                "fecha_verificacion": r.fecha_verificacion.isoformat() if r.fecha_verificacion else None,
                "encuestador": r.username,
                "latitud": r.latitud,
                "longitud": r.longitud
            })
            
        return {
            "success": True, "total": total, "page": page, "per_page": per_page,
            "medicos": medicos
        }
    except Exception as e:
        logger.error(f"[cliente_encuestador] Error cargando tabla médicos: {e}")
        return {
            "success": True, "total": 0, "page": page, "per_page": per_page,
            "medicos": []
        }


@router.get("/export")
async def api_export_excel(request: Request, db: AsyncSession = Depends(get_async_db), current_user: User = Depends(get_current_user)):
    await check_rol_cliente_encuestador(current_user, db)

    base_q, pc = get_base_query(db)
    base_q = apply_filters(base_q, request, pc)

    # Deduplicar: solo la encuesta más reciente por médico
    dedup = _latest_medico_centro_subq(db)
    base_q = base_q.filter(MedicoCentroEncuesta.id_medico_centro.in_(dedup))

    rows = (await db.execute(base_q.with_only_columns(
        Medico.id_medico_externo, Medico.apellido1, Medico.apellido2, Medico.nombre1, Medico.nombre2,
        Medico.especialidad, Medico.sub_especialidad, Medico.universidad_graduacion,
        Medico.ciudad, Medico.estado, Medico.telefono, Medico.whatsapp, Medico.email,
        Medico.linkedin, Medico.instagram,
        CentroSalud.nombre_centro, pc.c.nombre_clinica, pc.c.piso_consultorio, pc.c.direccion_especifica,
        pc.c.valor_consulta_rango, pc.c.promedio_pacientes_semanal_rango, pc.c.horarios_json,
        EncuestaCentro.fecha_verificacion, EncuestaCentro.fuente_informacion, User.username,
    ).order_by(desc(EncuestaCentro.fecha_verificacion), Medico.apellido1))).fetchall()

    columnas = [
        "ID Externo", "Apellidos", "Nombres", "Especialidad", "Sub-especialidad", "Universidad",
        "Ciudad", "Estado", "Teléfono", "WhatsApp", "Email", "LinkedIn", "Instagram",
        "Centro de Salud", "Consultorio", "Piso/Consultorio", "Dirección",
        "Valor Consulta", "Pacientes/Semana", "Días de Consulta",
        "Fecha Verificación", "Fuente", "Encuestador",
    ]
    data = []
    for r in rows:
        data.append({
            "ID Externo": r.id_medico_externo,
            "Apellidos": f"{r.apellido1} {r.apellido2 or ''}".strip(),
            "Nombres": f"{r.nombre1} {r.nombre2 or ''}".strip(),
            "Especialidad": r.especialidad,
            "Sub-especialidad": r.sub_especialidad,
            "Universidad": r.universidad_graduacion,
            "Ciudad": r.ciudad,
            "Estado": r.estado,
            "Teléfono": r.telefono,
            "WhatsApp": r.whatsapp,
            "Email": r.email,
            "LinkedIn": r.linkedin,
            "Instagram": r.instagram,
            "Centro de Salud": r.nombre_centro,
            "Consultorio": r.nombre_clinica,
            "Piso/Consultorio": r.piso_consultorio,
            "Dirección": r.direccion_especifica,
            "Valor Consulta": r.valor_consulta_rango,
            "Pacientes/Semana": r.promedio_pacientes_semanal_rango,
            "Días de Consulta": _dias_activos_str(r.horarios_json),
            "Fecha Verificación": r.fecha_verificacion.isoformat() if r.fecha_verificacion else None,
            "Fuente": r.fuente_informacion,
            "Encuestador": r.username,
        })

    df = pd.DataFrame(data, columns=columnas) if data else pd.DataFrame(columns=columnas)

    # engine='openpyxl' -- es el que está declarado en requirements.txt.
    # xlsxwriter (usado en reporteria.py) no lo está; ese endpoint probablemente
    # ya falla en producción con ModuleNotFoundError, pero no es parte de este cambio.
    from openpyxl.styles import Font, PatternFill

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Médicos')
        worksheet = writer.sheets['Médicos']
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='21262D', end_color='21262D', fill_type='solid')
        for col_num, value in enumerate(df.columns.values, start=1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            max_len = max(df[value].astype(str).map(len).max() if not df.empty else 0, len(str(value))) + 2
            worksheet.column_dimensions[cell.column_letter].width = min(max_len, 45)
    output.seek(0)

    filename = f"IQVIA_Medicos_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        output,
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
