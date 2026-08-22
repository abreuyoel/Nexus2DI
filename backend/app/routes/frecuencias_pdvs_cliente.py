"""CRUD de FRECUENCIAS_PDVS_CLIENTE: cuantas veces por semana debe visitarse
un PDV para un cliente dado."""
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Query, File, UploadFile, Form
import io
from openpyxl import load_workbook
from sqlalchemy import text, select, delete as sa_delete
from sqlalchemy.ext.asyncio import AsyncSession
from typing import List, Optional

from app.db.session import get_async_db
from app.core.dependencies import require_permission
from app.models.user import Usuario
from app.models.cliente import Cliente
from app.models.punto import PuntoInteres
from app.models.frecuencia_pdv_cliente import FrecuenciaPdvCliente
from app.schemas.frecuencia_pdv_cliente import (
    FrecuenciaPdvClienteCreate, FrecuenciaPdvClienteUpdate, FrecuenciaPdvClienteResponse,
    FrecuenciaBulkCreate,
)

router = APIRouter(prefix="/api/frecuencias-pdvs-cliente", tags=["Frecuencias PDVs Cliente"])


def _build_stmt():
    return (
        select(
            FrecuenciaPdvCliente,
            Cliente.nombre.label("cliente_nombre"),
            PuntoInteres.nombre.label("pdv_nombre"),
            Usuario.username.label("usuario_username")
        )
        .outerjoin(Cliente, Cliente.id == FrecuenciaPdvCliente.id_cliente)
        .outerjoin(PuntoInteres, PuntoInteres.id == FrecuenciaPdvCliente.id_punto_interes)
        .outerjoin(Usuario, Usuario.id == FrecuenciaPdvCliente.id_usuario)
    )


def _scope_analista_stmt(stmt, current_user: Usuario):
    if not (current_user.is_analyst and current_user.id_perfil):
        return stmt
    return stmt.filter(text("""
        EXISTS (
            SELECT 1 FROM RUTA_PROGRAMACION rp_a
            JOIN analistas_rutas ar_a ON rp_a.id_ruta = ar_a.id_ruta
            WHERE rp_a.id_punto_interes = FRECUENCIAS_PDVS_CLIENTE.id_punto_interes
              AND rp_a.id_cliente = FRECUENCIAS_PDVS_CLIENTE.id_cliente
              AND rp_a.activa = 1 AND ar_a.id_analista = :analista_id
        )
    """)).params(analista_id=int(current_user.id_perfil))


def _to_resp(row) -> FrecuenciaPdvClienteResponse:
    f = row.FrecuenciaPdvCliente
    return FrecuenciaPdvClienteResponse(
        id=f.id, id_cliente=f.id_cliente, id_punto_interes=f.id_punto_interes,
        frecuencia_semanal=float(f.frecuencia_semanal), observaciones=f.observaciones, activo=f.activo,
        fecha_creacion=f.fecha_creacion, fecha_modificacion=f.fecha_modificacion, id_usuario=f.id_usuario,
        cliente_nombre=row.cliente_nombre, pdv_nombre=row.pdv_nombre, usuario_username=row.usuario_username,
    )


@router.get("", response_model=List[FrecuenciaPdvClienteResponse])
async def list_frecuencias(
    id_cliente: Optional[int] = Query(None),
    id_punto_interes: Optional[str] = Query(None),
    activo: Optional[bool] = Query(None),
    db: AsyncSession = Depends(get_async_db),
    current_user: Usuario = Depends(require_permission('frecuencias-pdvs-cliente', 'read')),
):
    stmt = _build_stmt()
    if id_cliente is not None:
        stmt = stmt.filter(FrecuenciaPdvCliente.id_cliente == id_cliente)
    if id_punto_interes is not None:
        stmt = stmt.filter(FrecuenciaPdvCliente.id_punto_interes == id_punto_interes)
    if activo is not None:
        stmt = stmt.filter(FrecuenciaPdvCliente.activo == activo)
    stmt = _scope_analista_stmt(stmt, current_user)
    stmt = stmt.order_by(FrecuenciaPdvCliente.id.desc())
    rows = (await db.execute(stmt)).all()
    return [_to_resp(row) for row in rows]


@router.get("/pdvs-disponibles/{id_cliente}")
async def pdvs_disponibles_cliente(
    id_cliente: int,
    db: AsyncSession = Depends(get_async_db),
    current_user: Usuario = Depends(require_permission('frecuencias-pdvs-cliente', 'read')),
):
    """PDVs unicos donde aparece el cliente en RUTA_PROGRAMACION, marcando la
    frecuencia ya asignada (si existe) para poder editarla en la carga masiva.
    Si el que pregunta es analista, solo ve los PDVs de ESE cliente que caen
    dentro de sus propias rutas asignadas (analistas_rutas).
    Optimizado: una sola query con LEFT JOIN en vez de 2 queries separadas."""
    if not (await db.execute(select(Cliente).filter(Cliente.id == id_cliente))).scalars().first():
        raise HTTPException(404, "Cliente no existe")
    scope_sql = ""
    params = {"cid": id_cliente}
    if current_user.is_analyst and current_user.id_perfil:
        scope_sql = """
            AND EXISTS (SELECT 1 FROM analistas_rutas ar_a
                WHERE ar_a.id_ruta = rp.id_ruta AND ar_a.id_analista = :analista_id)
        """
        params["analista_id"] = int(current_user.id_perfil)
    rows = (await db.execute(text(f"""
        SELECT DISTINCT
            rp.id_punto_interes,
            rp.punto_interes,
            f.id_frecuencia_pdv_cliente,
            f.frecuencia_semanal,
            f.observaciones
        FROM RUTA_PROGRAMACION rp
        LEFT JOIN FRECUENCIAS_PDVS_CLIENTE f
            ON f.id_punto_interes = rp.id_punto_interes
            AND f.id_cliente = rp.id_cliente
        WHERE rp.id_cliente = :cid AND rp.activa = 1 AND rp.id_punto_interes IS NOT NULL
        {scope_sql}
        ORDER BY rp.punto_interes
    """), params)).fetchall()
    return [
        {
            "id_punto_interes": r.id_punto_interes,
            "pdv_nombre": r.punto_interes,
            "id_frecuencia": r.id_frecuencia_pdv_cliente,
            "frecuencia_semanal": float(r.frecuencia_semanal) if r.frecuencia_semanal is not None else None,
            "observaciones": r.observaciones,
        }
        for r in rows
    ]


@router.post("/bulk")
async def bulk_upsert_frecuencias(
    data: FrecuenciaBulkCreate,
    db: AsyncSession = Depends(get_async_db),
    current_user: Usuario = Depends(require_permission('frecuencias-pdvs-cliente.carga_masiva', 'read')),
):
    """Crea o actualiza (upsert) varias frecuencias de una vez para un mismo
    cliente — pensado para la carga masiva desde los PDVs de su programación."""
    if not (await db.execute(select(Cliente).filter(Cliente.id == data.id_cliente))).scalars().first():
        raise HTTPException(404, "Cliente no existe")
    creados = 0
    actualizados = 0
    for item in data.items:
        existente = (await db.execute(
            select(FrecuenciaPdvCliente).filter_by(
                id_cliente=data.id_cliente, id_punto_interes=item.id_punto_interes
            )
        )).scalars().first()
        if existente:
            existente.frecuencia_semanal = item.frecuencia_semanal
            existente.observaciones = item.observaciones
            existente.activo = True
            existente.id_usuario = current_user.id
            existente.fecha_modificacion = datetime.utcnow()
            actualizados += 1
        else:
            db.add(FrecuenciaPdvCliente(
                id_cliente=data.id_cliente, id_punto_interes=item.id_punto_interes,
                frecuencia_semanal=item.frecuencia_semanal, observaciones=item.observaciones,
                activo=True, id_usuario=current_user.id,
            ))
            creados += 1
    await db.commit()
    return {"creados": creados, "actualizados": actualizados}


@router.get("/{id_frecuencia}", response_model=FrecuenciaPdvClienteResponse)
async def get_frecuencia(id_frecuencia: int, db: AsyncSession = Depends(get_async_db), _: Usuario = Depends(require_permission('frecuencias-pdvs-cliente', 'read'))):
    row = (await db.execute(_build_stmt().filter(FrecuenciaPdvCliente.id == id_frecuencia))).first()
    if not row:
        raise HTTPException(404, "Registro no encontrado")
    return _to_resp(row)


@router.post("", response_model=FrecuenciaPdvClienteResponse, status_code=201)
async def create_frecuencia(
    data: FrecuenciaPdvClienteCreate,
    db: AsyncSession = Depends(get_async_db),
    current_user: Usuario = Depends(require_permission('frecuencias-pdvs-cliente.crear', 'read')),
):
    if not (await db.execute(select(Cliente).filter(Cliente.id == data.id_cliente))).scalars().first():
        raise HTTPException(404, "Cliente no existe")
    if not (await db.execute(select(PuntoInteres).filter(PuntoInteres.id == data.id_punto_interes))).scalars().first():
        raise HTTPException(404, "PDV no existe")
    f = FrecuenciaPdvCliente(
        id_cliente=data.id_cliente, id_punto_interes=data.id_punto_interes,
        frecuencia_semanal=data.frecuencia_semanal, observaciones=data.observaciones,
        activo=data.activo, id_usuario=current_user.id,
    )
    db.add(f)
    await db.commit()
    await db.refresh(f)
    return await get_frecuencia(f.id, db, current_user)


@router.put("/{id_frecuencia}", response_model=FrecuenciaPdvClienteResponse)
async def update_frecuencia(
    id_frecuencia: int,
    data: FrecuenciaPdvClienteUpdate,
    db: AsyncSession = Depends(get_async_db),
    current_user: Usuario = Depends(require_permission('frecuencias-pdvs-cliente.editar', 'read')),
):
    f = (await db.execute(select(FrecuenciaPdvCliente).filter(FrecuenciaPdvCliente.id == id_frecuencia))).scalars().first()
    if not f:
        raise HTTPException(404, "Registro no encontrado")
    if data.id_cliente is not None and not (await db.execute(select(Cliente).filter(Cliente.id == data.id_cliente))).scalars().first():
        raise HTTPException(404, "Cliente no existe")
    if data.id_punto_interes is not None and not (await db.execute(select(PuntoInteres).filter(PuntoInteres.id == data.id_punto_interes))).scalars().first():
        raise HTTPException(404, "PDV no existe")
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(f, k, v)
    f.id_usuario = current_user.id
    f.fecha_modificacion = datetime.utcnow()
    await db.commit()
    return await get_frecuencia(id_frecuencia, db, current_user)


@router.delete("/{id_frecuencia}")
async def delete_frecuencia(
    id_frecuencia: int,
    db: AsyncSession = Depends(get_async_db),
    _: Usuario = Depends(require_permission('frecuencias-pdvs-cliente.eliminar', 'read')),
):
    f = (await db.execute(select(FrecuenciaPdvCliente).filter(FrecuenciaPdvCliente.id == id_frecuencia))).scalars().first()
    if not f:
        raise HTTPException(404, "Registro no encontrado")
    await db.delete(f)
    await db.commit()
    return {"detail": "Registro eliminado"}


@router.post("/importar-excel")
async def importar_excel_frecuencias(
    id_cliente: int = Form(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_async_db),
    current_user: Usuario = Depends(require_permission('frecuencias-pdvs-cliente.carga_masiva', 'read')),
):
    """
    Recibe el archivo Excel cargado por el usuario, valida que corresponda al cliente,
    y realiza el upsert de frecuencias de forma masiva y optimizada en la base de datos.

    Optimizado: usa openpyxl read_only (no pandas) + bulk SQL (no N+1 queries).
    """
    import logging
    log = logging.getLogger("app")

    if not (await db.execute(select(Cliente).filter(Cliente.id == id_cliente))).scalars().first():
        raise HTTPException(404, "Cliente no existe")

    try:
        # ── 1. Leer archivo de forma async ──────────────────────────────
        contents = await file.read()

        # ── 2. Parsear con openpyxl read_only (ligero, sin pandas) ──────
        wb = load_workbook(io.BytesIO(contents), read_only=True, data_only=True)

        if "Frecuencias" not in wb.sheetnames:
            wb.close()
            raise HTTPException(400, "No se encontró la hoja 'Frecuencias' en el archivo.")

        ws = wb["Frecuencias"]

        # ── 3. Validar metadatos (filas 1-8) ────────────────────────────
        meta_rows = []
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, max_col=2, values_only=True), 1):
            meta_rows.append(row)
            if i >= 8:
                break

        if len(meta_rows) < 5 or len(meta_rows[4]) < 2:
            wb.close()
            raise HTTPException(400, "Formato de cabecera inválido en la hoja Frecuencias.")

        id_cliente_excel_raw = meta_rows[4][1]  # Fila 5 (índice 4), Columna B (índice 1)
        try:
            id_cliente_excel = int(id_cliente_excel_raw)
        except (ValueError, TypeError):
            wb.close()
            raise HTTPException(
                400,
                f"No se pudo leer el ID del cliente en la celda B5 del archivo. Encontrado: {id_cliente_excel_raw}"
            )

        if id_cliente_excel != id_cliente:
            wb.close()
            raise HTTPException(
                400,
                f"El cliente del archivo ({id_cliente_excel}) no coincide con el seleccionado ({id_cliente})."
            )

        # ── 4. Extraer datos (desde fila 10, saltando cabecera fila 9) ──
        items = []
        for row in ws.iter_rows(min_row=10, values_only=True):
            # openpyxl read_only puede devolver tuplas más cortas si las
            # últimas celdas están vacías — rellenar a 5 columnas mínimo
            cells = tuple(row) + (None,) * max(0, 5 - len(row)) if row else (None,) * 5

            id_pdv_raw = cells[0]
            if id_pdv_raw is None:
                continue
            id_pdv = str(id_pdv_raw).strip()
            if not id_pdv or id_pdv.lower() == "none":
                continue

            freq_raw = cells[3]  # Columna D: Frecuencia Semanal
            if freq_raw is None or str(freq_raw).strip() == "":
                continue

            try:
                freq_str = str(freq_raw).replace(",", ".")
                freq_val = float(freq_str)
            except (ValueError, TypeError):
                continue

            obs_raw = cells[4]  # Columna E: Observaciones
            obs_val = str(obs_raw).strip() if obs_raw is not None and str(obs_raw).strip() else None

            items.append({
                "id_pdv": id_pdv,
                "freq": freq_val,
                "obs": obs_val,
            })

        wb.close()

        if not items:
            raise HTTPException(400, "El archivo no contiene registros válidos para procesar.")

        # ── 5. Bulk upsert via temp table — todo server-side ───────────
        now = datetime.utcnow()

        # 5a. Crear tabla temporal y poblarla con los items del Excel
        await db.execute(text("""
            CREATE TABLE #freq_import (
                id_pdv VARCHAR(50) COLLATE DATABASE_DEFAULT,
                freq   NUMERIC(5,2),
                obs    VARCHAR(500) COLLATE DATABASE_DEFAULT
            )
        """))
        await db.execute(
            text("INSERT INTO #freq_import (id_pdv, freq, obs) VALUES (:id_pdv, :freq, :obs)"),
            items,
        )

        # 5b. UPDATE existentes (1 sola query, join server-side)
        upd_result = await db.execute(text("""
            UPDATE f
            SET f.frecuencia_semanal  = t.freq,
                f.observaciones       = t.obs,
                f.activo              = 1,
                f.id_usuario          = :uid,
                f.fecha_modificacion  = :now
            FROM FRECUENCIAS_PDVS_CLIENTE f
            INNER JOIN #freq_import t ON f.id_punto_interes = t.id_pdv
            WHERE f.id_cliente = :cid
        """), {"uid": current_user.id, "now": now, "cid": id_cliente})
        actualizados = upd_result.rowcount

        # 5c. INSERT nuevos (solo PDVs que existen en PUNTOS_INTERES1, 1 query)
        ins_result = await db.execute(text("""
            INSERT INTO FRECUENCIAS_PDVS_CLIENTE
                (id_cliente, id_punto_interes, frecuencia_semanal, observaciones, activo, id_usuario)
            SELECT :cid, t.id_pdv, t.freq, t.obs, 1, :uid
            FROM #freq_import t
            INNER JOIN PUNTOS_INTERES1 p ON p.identificador = t.id_pdv
            WHERE NOT EXISTS (
                SELECT 1 FROM FRECUENCIAS_PDVS_CLIENTE f
                WHERE f.id_cliente = :cid AND f.id_punto_interes = t.id_pdv
            )
        """), {"cid": id_cliente, "uid": current_user.id})
        creados = ins_result.rowcount

        # 5d. Obtener omitidos (PDVs del Excel que no existen en catálogo)
        omitidos_rows = (await db.execute(text("""
            SELECT t.id_pdv, t.freq, t.obs
            FROM #freq_import t
            WHERE NOT EXISTS (
                SELECT 1 FROM PUNTOS_INTERES1 p WHERE p.identificador = t.id_pdv
            )
            AND NOT EXISTS (
                SELECT 1 FROM FRECUENCIAS_PDVS_CLIENTE f
                WHERE f.id_cliente = :cid AND f.id_punto_interes = t.id_pdv
            )
        """), {"cid": id_cliente})).fetchall()

        omitidos = [
            {
                "id_punto_interes": r[0],
                "frecuencia_semanal": float(r[1]) if r[1] is not None else 0,
                "observaciones": r[2],
                "razon": "PDV no existe en el catálogo de Puntos de Interés",
            }
            for r in omitidos_rows
        ]

        await db.execute(text("DROP TABLE #freq_import"))
        await db.commit()

        result: dict = {"creados": creados, "actualizados": actualizados}
        if omitidos:
            result["omitidos"] = omitidos
        return result

    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        log.exception("importar-excel falló: %s", e)
        raise HTTPException(500, f"Error al procesar el archivo Excel en el servidor: {str(e)}")


