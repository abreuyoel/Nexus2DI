import json
import io
import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, desc, func, text, over as sql_over
from sqlalchemy.exc import IntegrityError
from typing import List, Optional, Any
from datetime import date, datetime, timedelta

from app.db.session import get_db
from app.core.dependencies import get_current_user
from app.modules.auth.entities import Usuario as User
from app.modules.surveyors.entities import JornadaEncuestador, CentroSalud, EncuestaCentro, Medico, MedicoCentroEncuesta, MedicoConsultorio
from app.modules.surveyors.dto import JornadaActivarRequest, CentroSaludCreate, EncuestaCentroCreate, MedicoCentroCreate
from app.modules.customer_service.entities import Solicitud

router = APIRouter(tags=["Encuestador"])


# ════════════════════════════════════════════════════════════════════════════
# 1. Rutas del Encuestador (antes routes/encuestador.py)
# ════════════════════════════════════════════════════════════════════════════

def check_rol_encuestador(current_user: User):
    # 12 = Encuestador (trabajo de campo). 13 = Cliente Encuestador/IQVIA --
    # además de ver el BI, puede entrar a activar jornadas y cargar médicos
    # como cualquier encuestador (decisión explícita, no es un descuido).
    if current_user.id_rol not in (12, 13) and not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Acceso denegado. Solo para Encuestadores.")


@router.get("/api/encuestador/jornada-activa")
def api_jornada_activa(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    check_rol_encuestador(current_user)
    
    jornada = db.query(JornadaEncuestador).filter(
        JornadaEncuestador.id_usuario == current_user.id,
        JornadaEncuestador.estado == 'En Progreso'
    ).order_by(desc(JornadaEncuestador.id_jornada)).first()
    
    if not jornada:
        return {"success": True, "activa": False}
        
    medicos_registrados = db.query(func.count(MedicoCentroEncuesta.id_medico_centro)).join(
        EncuestaCentro, EncuestaCentro.id_encuesta == MedicoCentroEncuesta.id_encuesta
    ).filter(EncuestaCentro.id_jornada == jornada.id_jornada).scalar() or 0
    
    centros_visitados = db.query(func.count(func.distinct(EncuestaCentro.id_centro))).filter(
        EncuestaCentro.id_jornada == jornada.id_jornada
    ).scalar() or 0
    
    return {
        "success": True,
        "activa": True,
        "id_jornada": jornada.id_jornada,
        "fecha_inicio": jornada.fecha_inicio.isoformat() if jornada.fecha_inicio else None,
        "ciudad": jornada.ciudad,
        "estado_geo": jornada.estado_geo,
        "medicos_registrados": medicos_registrados,
        "centros_visitados": centros_visitados
    }


@router.post("/api/encuestador/activar-jornada")
def api_activar_jornada(req: JornadaActivarRequest, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    check_rol_encuestador(current_user)
    
    existente = db.query(JornadaEncuestador).filter(
        JornadaEncuestador.id_usuario == current_user.id,
        JornadaEncuestador.estado == 'En Progreso'
    ).first()
    if existente:
        return {"success": True, "id_jornada": existente.id_jornada, "ya_activa": True}
        
    nueva_jornada = JornadaEncuestador(
        id_usuario=current_user.id,
        estado='En Progreso',
        latitud=req.latitud,
        longitud=req.longitud,
        ciudad=req.ciudad.strip() if req.ciudad else None,
        estado_geo=req.estado_geo.strip() if req.estado_geo else None
    )
    db.add(nueva_jornada)
    db.commit()
    db.refresh(nueva_jornada)
    
    return {
        "success": True, 
        "id_jornada": nueva_jornada.id_jornada, 
        "fecha_inicio": nueva_jornada.fecha_inicio.isoformat() if nueva_jornada.fecha_inicio else None
    }


@router.post("/api/encuestador/finalizar-jornada")
def api_finalizar_jornada(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    check_rol_encuestador(current_user)
    
    jornada = db.query(JornadaEncuestador).filter(
        JornadaEncuestador.id_usuario == current_user.id,
        JornadaEncuestador.estado == 'En Progreso'
    ).first()
    
    if jornada:
        encuestas_abiertas = db.query(EncuestaCentro).filter(
            EncuestaCentro.id_jornada == jornada.id_jornada,
            EncuestaCentro.estado == 'Abierta'
        ).all()
        for e in encuestas_abiertas:
            e.estado = 'Cerrada'
            
        jornada.estado = 'Finalizada'
        jornada.fecha_fin = datetime.utcnow()
        db.commit()
        
    return {"success": True, "message": "Jornada finalizada"}


@router.get("/api/encuestador/centros")
def api_centros_list(q: str = "", db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    check_rol_encuestador(current_user)
    
    query = db.query(CentroSalud)
    if q.strip():
        search = f"%{q.strip()}%"
        query = query.filter(
            or_(
                CentroSalud.nombre_centro.ilike(search),
                CentroSalud.ciudad.ilike(search),
                CentroSalud.estado.ilike(search)
            )
        )
    centros = query.order_by(CentroSalud.nombre_centro).limit(50).all()
    
    return {
        "success": True,
        "centros": [
            {
                "id_centro": c.id_centro,
                "nombre_centro": c.nombre_centro,
                "direccion_completa": c.direccion_completa,
                "ciudad": c.ciudad,
                "estado": c.estado
            } for c in centros
        ]
    }


@router.post("/api/encuestador/centros")
def api_centros_create(req: CentroSaludCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    check_rol_encuestador(current_user)
    
    datos_centro = {
        "nombre_centro": req.nombre_centro.strip(),
        "direccion_completa": req.direccion_completa.strip(),
        "ciudad": req.ciudad.strip() if req.ciudad else None,
        "estado": req.estado.strip() if req.estado else None
    }
    
    # Prevenir duplicados si la cola offline reintenta el POST por timeout
    search_str = f'%"nombre_centro": "{req.nombre_centro.strip()}"%'
    existente = db.query(Solicitud).filter(
        Solicitud.user_id == current_user.id,
        Solicitud.tipo == "creacion_centro_salud",
        Solicitud.estado == "pendiente",
        Solicitud.descripcion.like(search_str)
    ).first()
    
    if existente:
        return {
            "success": True,
            "solicitud_id": existente.id,
            "message": "Solicitud ya estaba registrada (reintento de cola offline)."
        }

    nueva_solicitud = Solicitud(
        user_id=current_user.id,
        tipo="creacion_centro_salud",
        descripcion=json.dumps(datos_centro),
        estado="pendiente"
    )
    db.add(nueva_solicitud)
    db.commit()
    db.refresh(nueva_solicitud)
    
    return {
        "success": True,
        "solicitud_id": nueva_solicitud.id,
        "message": "Solicitud de creación de centro enviada a Atención al Cliente para su aprobación."
    }


@router.get("/api/encuestador/encuesta-abierta")
def api_encuesta_abierta(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    check_rol_encuestador(current_user)
    
    jornada = db.query(JornadaEncuestador).filter(
        JornadaEncuestador.id_usuario == current_user.id,
        JornadaEncuestador.estado == 'En Progreso'
    ).first()
    
    if not jornada:
        return {"success": True, "tiene_encuesta": False, "jornada_activa": False}
        
    encuesta = db.query(EncuestaCentro, CentroSalud).join(
        CentroSalud, CentroSalud.id_centro == EncuestaCentro.id_centro
    ).filter(
        EncuestaCentro.id_jornada == jornada.id_jornada,
        EncuestaCentro.estado == 'Abierta'
    ).order_by(desc(EncuestaCentro.id_encuesta)).first()
    
    if not encuesta:
        return {"success": True, "tiene_encuesta": False, "jornada_activa": True, "id_jornada": jornada.id_jornada}
        
    ec, cs = encuesta
    
    medicos_cargados = db.query(Medico, MedicoCentroEncuesta).join(
        MedicoCentroEncuesta, MedicoCentroEncuesta.id_medico == Medico.id_medico
    ).filter(
        MedicoCentroEncuesta.id_encuesta == ec.id_encuesta
    ).order_by(desc(MedicoCentroEncuesta.id_medico_centro)).all()
    
    medicos_resp = []
    for m, mce in medicos_cargados:
        first_consultorio = db.query(MedicoConsultorio).filter(MedicoConsultorio.id_medico == m.id_medico).first()
        val = first_consultorio.valor_consulta_rango if first_consultorio else 'N/A'
        pacs = first_consultorio.promedio_pacientes_semanal_rango if first_consultorio else 'N/A'
        medicos_resp.append({
            "id_medico_centro": mce.id_medico_centro,
            "id_medico": m.id_medico,
            "id_medico_externo": m.id_medico_externo,
            "apellido1": m.apellido1,
            "apellido2": m.apellido2,
            "nombre1": m.nombre1,
            "nombre2": m.nombre2,
            "especialidad": m.especialidad,
            "valor_consulta_rango": val,
            "promedio_pacientes_semanal_rango": pacs
        })
    
    return {
        "success": True,
        "tiene_encuesta": True,
        "jornada_activa": True,
        "id_jornada": jornada.id_jornada,
        "id_encuesta": ec.id_encuesta,
        "id_centro": cs.id_centro,
        "nombre_centro": cs.nombre_centro,
        "ciudad": cs.ciudad,
        "estado": cs.estado,
        "fecha_verificacion": ec.fecha_verificacion.isoformat() if ec.fecha_verificacion else None,
        "fuente_informacion": ec.fuente_informacion,
        "medicos": medicos_resp
    }


@router.post("/api/encuestador/encuestas")
def api_encuestas_crear(req: EncuestaCentroCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    check_rol_encuestador(current_user)
    
    jornada = db.query(JornadaEncuestador).filter(
        JornadaEncuestador.id_usuario == current_user.id,
        JornadaEncuestador.estado == 'En Progreso'
    ).first()
    
    if not jornada:
        raise HTTPException(status_code=400, detail="Debes activar una jornada primero")
        
    existente = db.query(EncuestaCentro).filter(
        EncuestaCentro.id_jornada == jornada.id_jornada,
        EncuestaCentro.estado == 'Abierta'
    ).first()
    
    if existente:
        if existente.id_centro == req.id_centro:
            # Si es el mismo centro, es un reintento de la cola offline que falló por timeout
            # pero sí llegó a crearse en el backend. Devolver éxito para destrabar la cola.
            return {"success": True, "id_encuesta": existente.id_encuesta, "id_jornada": jornada.id_jornada}
        raise HTTPException(status_code=409, detail=f"Ya tienes una encuesta abierta. Ciérrala antes de iniciar otra (ID {existente.id_encuesta}).")
        
    nueva_encuesta = EncuestaCentro(
        id_usuario=current_user.id,
        id_centro=req.id_centro,
        id_jornada=jornada.id_jornada,
        fecha_verificacion=datetime.utcnow().date(),
        fuente_informacion=req.fuente_informacion,
        notas_generales=req.notas_generales,
        estado='Abierta'
    )
    db.add(nueva_encuesta)
    # Mismo riesgo de carrera que medico-centro (ver comentario ahí): el
    # SELECT de "existente" de arriba no es atómico con el INSERT. El índice
    # único filtrado UQ_encuesta_abierta_por_jornada (solo WHERE estado=
    # 'Abierta') es la protección real -- si se pierde la carrera, se
    # devuelve la encuesta que sí ganó en vez de un 500.
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        ganadora = db.query(EncuestaCentro).filter(
            EncuestaCentro.id_jornada == jornada.id_jornada,
            EncuestaCentro.estado == 'Abierta'
        ).first()
        if ganadora:
            return {"success": True, "id_encuesta": ganadora.id_encuesta, "id_jornada": jornada.id_jornada}
        raise HTTPException(status_code=409, detail="Ya tienes una encuesta abierta.")
    db.refresh(nueva_encuesta)
    
    return {"success": True, "id_encuesta": nueva_encuesta.id_encuesta, "id_jornada": jornada.id_jornada}


@router.post("/api/encuestador/encuestas/{id_encuesta}/cerrar")
def api_encuesta_cerrar(id_encuesta: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    check_rol_encuestador(current_user)
    
    encuesta = db.query(EncuestaCentro).filter(
        EncuestaCentro.id_encuesta == id_encuesta,
        EncuestaCentro.id_usuario == current_user.id
    ).first()
    
    if encuesta:
        encuesta.estado = 'Cerrada'
        db.commit()
        
    return {"success": True}


@router.get("/api/encuestador/medicos/buscar")
def api_medicos_buscar(q: str = "", db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    check_rol_encuestador(current_user)
    
    if not q.strip():
        return {"success": True, "medicos": []}
        
    search = f"%{q.strip()}%"
    medicos = db.query(Medico).filter(
        or_(
            Medico.id_medico_externo.ilike(search),
            Medico.apellido1.ilike(search),
            Medico.apellido2.ilike(search),
            Medico.nombre1.ilike(search),
            Medico.nombre2.ilike(search)
        )
    ).order_by(Medico.apellido1, Medico.nombre1).limit(25).all()
    
    return {
        "success": True,
        "medicos": [
            {
                "id_medico": m.id_medico,
                "id_medico_externo": m.id_medico_externo,
                "apellido1": m.apellido1,
                "apellido2": m.apellido2,
                "nombre1": m.nombre1,
                "nombre2": m.nombre2,
                "especialidad": m.especialidad,
                "sub_especialidad": m.sub_especialidad,
                "universidad_graduacion": m.universidad_graduacion,
                "nro_MPPS": m.nro_MPPS,
                "nro_colegiado": m.nro_colegiado,
                "ciudad": m.ciudad,
                "estado": m.estado,
                "telefono": m.telefono,
                "whatsapp": m.whatsapp,
                "email": m.email,
                "linkedin": m.linkedin,
                "instagram": m.instagram
            } for m in medicos
        ]
    }


@router.post("/api/encuestador/medico-centro")
def api_medico_centro_save(req: MedicoCentroCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    check_rol_encuestador(current_user)
    
    jornada = db.query(JornadaEncuestador).filter(
        JornadaEncuestador.id_usuario == current_user.id,
        JornadaEncuestador.estado == 'En Progreso'
    ).first()
    if not jornada:
        raise HTTPException(status_code=400, detail="No tienes jornada activa")
        
    encuesta = db.query(EncuestaCentro).filter(
        EncuestaCentro.id_jornada == jornada.id_jornada,
        EncuestaCentro.estado == 'Abierta'
    ).first()
    if not encuesta:
        raise HTTPException(status_code=400, detail="No tienes encuesta abierta")
        
    id_medico = req.id_medico
    if not id_medico:
        if not req.apellido1 or not req.apellido2 or not req.nombre1 or not req.especialidad or not req.ciudad or not req.estado:
            raise HTTPException(status_code=400, detail="Faltan campos obligatorios del médico")
            
        existente = None
        if req.id_medico_externo and req.id_medico_externo != "000000" and req.id_medico_externo.strip() != "":
            existente = db.query(Medico).filter(Medico.id_medico_externo == req.id_medico_externo).first()
        else:
            # Si no hay cédula, intentamos evitar duplicar el mismo médico en reintentos offline de la cola
            # matcheando por nombre exacto.
            existente = db.query(Medico).filter(
                func.lower(Medico.nombre1) == req.nombre1.lower(),
                func.lower(Medico.apellido1) == req.apellido1.lower(),
                func.lower(Medico.especialidad) == req.especialidad.lower()
            ).order_by(desc(Medico.id_medico)).first()
            
        if existente:
            id_medico = existente.id_medico
        else:
            nuevo_medico = Medico(
                id_medico_externo=req.id_medico_externo,
                apellido1=req.apellido1,
                apellido2=req.apellido2,
                nombre1=req.nombre1,
                nombre2=req.nombre2,
                especialidad=req.especialidad,
                sub_especialidad=req.sub_especialidad,
                universidad_graduacion=req.universidad_graduacion,
                nro_MPPS=req.nro_MPPS,
                nro_colegiado=req.nro_colegiado,
                ciudad=req.ciudad,
                estado=req.estado,
                telefono=req.telefono,
                whatsapp=req.whatsapp,
                email=req.email,
                linkedin=req.linkedin,
                instagram=req.instagram
            )
            db.add(nuevo_medico)
            db.commit()
            db.refresh(nuevo_medico)
            id_medico = nuevo_medico.id_medico
            
    dup = db.query(MedicoCentroEncuesta).filter(
        MedicoCentroEncuesta.id_encuesta == encuesta.id_encuesta,
        MedicoCentroEncuesta.id_medico == id_medico
    ).first()
    
    if dup:
        raise HTTPException(status_code=409, detail="Este médico ya fue registrado en esta encuesta del centro.")
        
    m_c_e = MedicoCentroEncuesta(
        id_encuesta=encuesta.id_encuesta,
        id_medico=id_medico
    )
    db.add(m_c_e)
    
    for cons in req.consultorios:
        nuevo_consultorio = MedicoConsultorio(
            id_medico=id_medico,
            nombre_clinica=cons.nombre_clinica,
            piso_consultorio=cons.piso_consultorio,
            direccion_especifica=cons.direccion_especifica,
            horarios_json=cons.horarios_json,
            valor_consulta_rango=cons.valor_consulta_rango,
            promedio_pacientes_semanal_rango=cons.promedio_pacientes_semanal_rango
        )
        db.add(nuevo_consultorio)
        
    # El SELECT de arriba (dup) NO alcanza solo: con 2 réplicas del backend
    # corriendo en paralelo, dos requests casi simultáneas -- doble tap del
    # encuestador, o el reintento automático de la cola offline mientras el
    # primer intento seguía en vuelo con mala señal -- pueden pasar el
    # chequeo las dos antes de que ninguna haga commit. Es la causa
    # confirmada de médicos duplicados reportada en campo. El índice único
    # UQ_medico_centro_encuesta (ver migración SQL) es la protección real:
    # si se pierde la carrera acá, el commit falla con IntegrityError y se
    # trata como éxito (ya quedó registrado por el otro intento), en vez de
    # tirarle un 500 al encuestador -- y al ser una sola transacción, los
    # MedicoConsultorio de arriba se revierten con el rollback, así que
    # tampoco quedan consultorios duplicados sueltos.
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        cnt = db.query(func.count(MedicoCentroEncuesta.id_medico_centro)).filter(
            MedicoCentroEncuesta.id_encuesta == encuesta.id_encuesta
        ).scalar() or 0
        return {
            "success": True,
            "id_medico": id_medico,
            "id_encuesta": encuesta.id_encuesta,
            "medicos_en_centro": cnt,
            "ya_registrado": True,
        }
    
    cnt = db.query(func.count(MedicoCentroEncuesta.id_medico_centro)).filter(
        MedicoCentroEncuesta.id_encuesta == encuesta.id_encuesta
    ).scalar() or 0
    
    return {
        "success": True,
        "id_medico": id_medico,
        "id_encuesta": encuesta.id_encuesta,
        "medicos_en_centro": cnt
    }


@router.get("/medico/{id_medico}")
def api_medico_detalle(id_medico: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Datos completos de un médico ya registrado + todos sus consultorios,
    para precargar el formulario en modo edición. El médico es una entidad
    global (compartida entre todos los centros donde se lo haya registrado,
    matcheada por id_medico_externo) -- por eso acá aparecen TODOS sus
    consultorios, no solo los de la encuesta/centro desde donde se entró a
    editar. Es el mismo dato que ya se ve al buscar un médico existente en
    el formulario de alta, solo que ahora también se puede guardar."""
    check_rol_encuestador(current_user)

    medico = db.query(Medico).filter(Medico.id_medico == id_medico).first()
    if not medico:
        raise HTTPException(status_code=404, detail="Médico no encontrado")

    consultorios = db.query(MedicoConsultorio).filter(MedicoConsultorio.id_medico == id_medico).order_by(MedicoConsultorio.id_consultorio).all()

    return {
        "success": True,
        "id_medico": medico.id_medico,
        "id_medico_externo": medico.id_medico_externo,
        "apellido1": medico.apellido1,
        "apellido2": medico.apellido2,
        "nombre1": medico.nombre1,
        "nombre2": medico.nombre2,
        "especialidad": medico.especialidad,
        "sub_especialidad": medico.sub_especialidad,
        "universidad_graduacion": medico.universidad_graduacion,
        "nro_MPPS": medico.nro_MPPS,
        "nro_colegiado": medico.nro_colegiado,
        "ciudad": medico.ciudad,
        "estado": medico.estado,
        "telefono": medico.telefono,
        "whatsapp": medico.whatsapp,
        "email": medico.email,
        "linkedin": medico.linkedin,
        "instagram": medico.instagram,
        "consultorios": [
            {
                "id_consultorio": c.id_consultorio,
                "nombre_clinica": c.nombre_clinica,
                "piso_consultorio": c.piso_consultorio,
                "direccion_especifica": c.direccion_especifica,
                "valor_consulta_rango": c.valor_consulta_rango,
                "promedio_pacientes_semanal_rango": c.promedio_pacientes_semanal_rango,
                "horarios_json": c.horarios_json,
            } for c in consultorios
        ],
    }

@router.post("/medico/{id_medico}/editar")
def api_medico_editar(id_medico: int, req: MedicoCentroCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Edita los datos propios del médico y reemplaza por completo su lista
    de consultorios (delete-then-insert, mismo shape que el alta) -- más
    simple y menos propenso a errores que tratar de diffear cuál consultorio
    es cuál cuando el formulario no manda ids estables por fila."""
    check_rol_encuestador(current_user)

    medico = db.query(Medico).filter(Medico.id_medico == id_medico).first()
    if not medico:
        raise HTTPException(status_code=404, detail="Médico no encontrado")

    if not req.apellido1 or not req.apellido2 or not req.nombre1 or not req.especialidad or not req.ciudad or not req.estado:
        raise HTTPException(status_code=400, detail="Faltan campos obligatorios del médico")

    medico.apellido1 = req.apellido1
    medico.apellido2 = req.apellido2
    medico.nombre1 = req.nombre1
    medico.nombre2 = req.nombre2
    medico.especialidad = req.especialidad
    medico.sub_especialidad = req.sub_especialidad
    medico.universidad_graduacion = req.universidad_graduacion
    medico.nro_MPPS = req.nro_MPPS
    medico.nro_colegiado = req.nro_colegiado
    medico.ciudad = req.ciudad
    medico.estado = req.estado
    medico.telefono = req.telefono
    medico.whatsapp = req.whatsapp
    medico.email = req.email
    medico.linkedin = req.linkedin
    medico.instagram = req.instagram
    # id_medico_externo (cédula) NO se toca acá -- es la clave con la que se
    # matchea "¿ya existe este médico?" al buscarlo desde otro centro;
    # cambiarla desde acá podría duplicarlo en vez de identificarlo.

    db.query(MedicoConsultorio).filter(MedicoConsultorio.id_medico == id_medico).delete()
    for cons in req.consultorios:
        db.add(MedicoConsultorio(
            id_medico=id_medico,
            nombre_clinica=cons.nombre_clinica,
            piso_consultorio=cons.piso_consultorio,
            direccion_especifica=cons.direccion_especifica,
            horarios_json=cons.horarios_json,
            valor_consulta_rango=cons.valor_consulta_rango,
            promedio_pacientes_semanal_rango=cons.promedio_pacientes_semanal_rango,
        ))

    db.commit()
    return {"success": True, "id_medico": id_medico}


@router.get("/api/encuestador/catalogos")
def api_catalogos(current_user: User = Depends(get_current_user)):
    check_rol_encuestador(current_user)
    return {
        "valor_consulta_rangos": [
            "Menos de 30$", "Entre 30$ a 50$", "Entre 50$ a 60$",
            "Entre 60$ a 100$", "Más de 100$"
        ],
        "promedio_pacientes_rangos": [
            "1 a 5 pacientes", "6 a 10 pacientes", "11 a 15 pacientes",
            "16 a 20 pacientes", "21 a 30 pacientes", "Más de 30 pacientes"
        ],
        "fuentes_informacion": [
            "Visita presencial", "Llamada telefónica", "Referencia",
            "Página web del centro", "Redes sociales", "Otra"
        ],
        "dias_consulta": ["Lunes", "Martes", "Miércoles", "Jueves",
                          "Viernes", "Sábado", "Domingo"]
    }


# ════════════════════════════════════════════════════════════════════════════
# 2. Cliente Encuestador (antes routes/cliente_encuestador.py)
# ════════════════════════════════════════════════════════════════════════════

def check_rol_cliente_encuestador(current_user: User, db: Session):
    # id_rol=13 (IQVIA) y admin siempre pasan. Además, cualquier usuario con
    # el permiso 'cliente-encuestador' concedido a mano (Permisos -> por
    # usuario) también entra, sin importar su rol -- ej. un supervisor del
    # equipo de encuestadores (rol 12) al que se le da acceso de lectura al
    # BI sin volverlo IQVIA ni crear un rol nuevo solo para eso.
    # No se usa current_user.has_permission()/.permisos: esa relación es
    # lazy="noload" y get_current_user no la carga -- siempre saldría vacía.
    # Se consulta usuario_permisos directo, mismo patrón que require_permission()
    # en core/dependencies.py (el único camino que de verdad funciona hoy).
    if current_user.id_rol == 13 or current_user.is_admin:
        return
    perm = db.execute(text("""
        SELECT can_read FROM usuario_permisos WHERE id_usuario = :uid AND module = 'cliente-encuestador'
    """), {"uid": current_user.id}).fetchone()
    if perm and perm[0]:
        return
    raise HTTPException(status_code=403, detail="Acceso denegado. Solo para Cliente Encuestador.")


# Mismas keys abreviadas que medico-form.component.ts usa en el JSON de
# horarios (diasList) -- el filtro/gráfico de días tiene que buscar estas
# mismas keys en horarios_json o nunca va a matchear nada.
DIAS_ABREV = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]

def _dias_activos_str(horarios_json: Optional[str]) -> str:
    if not horarios_json:
        return ''
    try:
        h = json.loads(horarios_json)
    except (TypeError, ValueError):
        return ''
    return ', '.join(d for d in DIAS_ABREV if isinstance(h, dict) and h.get(d, {}).get('activo'))

def _primer_consultorio_alias(db: Session):
    """Un MedicoConsultorio por médico (el de menor id) para las métricas y
    filtros a nivel médico que antes vivían 1:1 en medico_centro_encuesta
    (valor_consulta_rango, promedio_pacientes_semanal_rango, días de
    consulta) -- con consultorios dinámicos un médico puede tener N, se usa
    el primero como representativo (mismo criterio ya usado en
    /encuesta-abierta para no inflar el conteo de médicos con joins 1:N)."""
    min_ids = db.query(
        MedicoConsultorio.id_medico.label('id_medico'),
        func.min(MedicoConsultorio.id_consultorio).label('min_id')
    ).group_by(MedicoConsultorio.id_medico).subquery()
    return db.query(MedicoConsultorio).join(
        min_ids, MedicoConsultorio.id_consultorio == min_ids.c.min_id
    ).subquery()


def _latest_medico_centro_subq(db: Session):
    """Subquery con los id_medico_centro de la encuesta más reciente por médico.
    ROW_NUMBER() particionado por id_medico, ordenado por fecha_verificacion DESC
    y id_medico_centro DESC como desempate. Evita médicos duplicados en el BI."""
    rn = db.query(
        MedicoCentroEncuesta.id_medico_centro,
        func.row_number().over(
            partition_by=Medico.id_medico,
            order_by=[desc(EncuestaCentro.fecha_verificacion), desc(MedicoCentroEncuesta.id_medico_centro)]
        ).label('rn')
    ).join(
        EncuestaCentro, EncuestaCentro.id_encuesta == MedicoCentroEncuesta.id_encuesta
    ).join(
        Medico, Medico.id_medico == MedicoCentroEncuesta.id_medico
    ).subquery()
    return db.query(rn.c.id_medico_centro).filter(rn.c.rn == 1).subquery()


def get_base_query(db: Session):
    pc = _primer_consultorio_alias(db)
    query = db.query(MedicoCentroEncuesta, EncuestaCentro, Medico, CentroSalud, User, JornadaEncuestador).join(
        EncuestaCentro, EncuestaCentro.id_encuesta == MedicoCentroEncuesta.id_encuesta
    ).join(
        Medico, Medico.id_medico == MedicoCentroEncuesta.id_medico
    ).join(
        CentroSalud, CentroSalud.id_centro == EncuestaCentro.id_centro
    ).join(
        User, User.id == EncuestaCentro.id_usuario
    ).outerjoin(
        JornadaEncuestador, JornadaEncuestador.id_jornada == EncuestaCentro.id_jornada
    ).outerjoin(
        pc, pc.c.id_medico == Medico.id_medico
    )
    return query, pc


def apply_filters(query, req: Request, pc):
    q_params = req.query_params
    
    fdesde = q_params.get("fecha_desde")
    fhasta = q_params.get("fecha_hasta")
    if fdesde: query = query.filter(EncuestaCentro.fecha_verificacion >= fdesde)
    if fhasta: query = query.filter(EncuestaCentro.fecha_verificacion <= fhasta)
    
    def apply_in(col, param_name):
        vals = q_params.getlist(param_name)
        if len(vals) == 1 and ',' in vals[0]:
            vals = [v.strip() for v in vals[0].split(',')]
        vals = [v for v in vals if v]
        if vals:
            return query.filter(col.in_(vals))
        return query
        
    query = apply_in(Medico.estado, "estados")
    query = apply_in(Medico.ciudad, "ciudades")
    query = apply_in(Medico.especialidad, "especialidades")
    query = apply_in(Medico.sub_especialidad, "sub_especialidades")
    query = apply_in(Medico.universidad_graduacion, "universidades")
    query = apply_in(CentroSalud.id_centro, "centros")
    query = apply_in(EncuestaCentro.id_usuario, "encuestadores")
    query = apply_in(EncuestaCentro.fuente_informacion, "fuentes")
    query = apply_in(pc.c.valor_consulta_rango, "valor_consulta_rangos")
    query = apply_in(pc.c.promedio_pacientes_semanal_rango, "promedio_pacientes_rangos")
    
    dias = q_params.getlist("dias_consulta")
    if len(dias) == 1 and ',' in dias[0]: dias = [d.strip() for d in dias[0].split(',')]
    dias = [d for d in dias if d]
    if dias:
        # horarios_json es {"Lun":{"activo":true,...}, ...} -- basta con
        # buscar la subcadena literal, sin necesidad de OPENJSON.
        ors = [pc.c.horarios_json.ilike(f'%"{d}":{{"activo":true%') for d in dias]
        query = query.filter(or_(*ors))
        
    return query


@router.get("/api/cliente-encuestador/filtros")
def api_filtros(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    check_rol_cliente_encuestador(current_user, db)
    
    especialidades = [r[0] for r in db.query(Medico.especialidad).distinct().filter(Medico.especialidad != None).order_by(Medico.especialidad).all()]
    sub_especialidades = [r[0] for r in db.query(Medico.sub_especialidad).distinct().filter(Medico.sub_especialidad != None).order_by(Medico.sub_especialidad).all()]
    estados = [r[0] for r in db.query(Medico.estado).distinct().filter(Medico.estado != None).order_by(Medico.estado).all()]
    ciudades = [r[0] for r in db.query(Medico.ciudad).distinct().filter(Medico.ciudad != None).order_by(Medico.ciudad).all()]
    universidades = [r[0] for r in db.query(Medico.universidad_graduacion).distinct().filter(Medico.universidad_graduacion != None).order_by(Medico.universidad_graduacion).all()]
    
    centros = [{"id_centro": r.id_centro, "nombre_centro": r.nombre_centro} for r in db.query(CentroSalud.id_centro, CentroSalud.nombre_centro).order_by(CentroSalud.nombre_centro).all()]
    encuestadores = [{"id_usuario": r.id, "username": r.username} for r in db.query(User.id, User.username).join(EncuestaCentro, EncuestaCentro.id_usuario == User.id).distinct().order_by(User.username).all()]
    
    fuentes = [r[0] for r in db.query(EncuestaCentro.fuente_informacion).distinct().filter(EncuestaCentro.fuente_informacion != None).order_by(EncuestaCentro.fuente_informacion).all()]
    valor_rangos = [r[0] for r in db.query(MedicoConsultorio.valor_consulta_rango).distinct().order_by(MedicoConsultorio.valor_consulta_rango).all()]
    pac_rangos = [r[0] for r in db.query(MedicoConsultorio.promedio_pacientes_semanal_rango).distinct().order_by(MedicoConsultorio.promedio_pacientes_semanal_rango).all()]
    
    return {
        "success": True,
        "especialidades": especialidades,
        "sub_especialidades": sub_especialidades,
        "estados": estados,
        "ciudades": ciudades,
        "universidades": universidades,
        "centros": centros,
        "encuestadores": encuestadores,
        "fuentes": fuentes,
        "valor_consulta_rangos": valor_rangos,
        "promedio_pacientes_rangos": pac_rangos,
        "dias_consulta": DIAS_ABREV
    }


@router.get("/api/cliente-encuestador/kpis")
def api_kpis(request: Request, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    check_rol_cliente_encuestador(current_user, db)
    q, pc = get_base_query(db)
    q = apply_filters(q, request, pc)
    
    total_medicos = q.with_entities(func.count(func.distinct(Medico.id_medico))).scalar() or 0
    total_centros = q.with_entities(func.count(func.distinct(CentroSalud.id_centro))).scalar() or 0
    total_especialidades = q.with_entities(func.count(func.distinct(Medico.especialidad))).scalar() or 0
    total_estados = q.with_entities(func.count(func.distinct(Medico.estado))).scalar() or 0
    total_ciudades = q.with_entities(func.count(func.distinct(Medico.ciudad))).scalar() or 0
    total_encuestas = q.with_entities(func.count(func.distinct(EncuestaCentro.id_encuesta))).scalar() or 0
    
    thirty_days_ago = datetime.utcnow().date() - timedelta(days=30)
    encuestas_30d = q.filter(EncuestaCentro.fecha_verificacion >= thirty_days_ago).with_entities(func.count(func.distinct(EncuestaCentro.id_encuesta))).scalar() or 0
    
    # "Tiene 2do consultorio" ya no es una columna -- es contar cuántos
    # médicos (del set ya filtrado) tienen más de 1 fila en medico_consultorios.
    medico_ids = [r[0] for r in q.with_entities(func.distinct(Medico.id_medico)).all()]
    dos_cons = 0
    if medico_ids:
        dos_cons = db.query(MedicoConsultorio.id_medico).filter(
            MedicoConsultorio.id_medico.in_(medico_ids)
        ).group_by(MedicoConsultorio.id_medico).having(func.count(MedicoConsultorio.id_consultorio) > 1).count()
    pct_dos = round((dos_cons * 100.0) / total_medicos, 1) if total_medicos else 0.0
    
    wa = q.filter(Medico.whatsapp != None, Medico.whatsapp != '').with_entities(func.count(func.distinct(Medico.id_medico))).scalar() or 0
    em = q.filter(Medico.email != None, Medico.email != '').with_entities(func.count(func.distinct(Medico.id_medico))).scalar() or 0
    tel = q.filter(Medico.telefono != None, Medico.telefono != '').with_entities(func.count(func.distinct(Medico.id_medico))).scalar() or 0
    ig = q.filter(Medico.instagram != None, Medico.instagram != '').with_entities(func.count(func.distinct(Medico.id_medico))).scalar() or 0
    li = q.filter(Medico.linkedin != None, Medico.linkedin != '').with_entities(func.count(func.distinct(Medico.id_medico))).scalar() or 0
    
    def pct(x): return round((x * 100.0) / total_medicos, 1) if total_medicos else 0.0
    
    # --- CHART DATA ---
    esp_data = q.with_entities(Medico.especialidad, func.count(func.distinct(Medico.id_medico))).group_by(Medico.especialidad).all()
    esp_chart = [{"name": r[0] or "N/A", "value": r[1]} for r in esp_data]

    est_data = q.with_entities(Medico.estado, func.count(func.distinct(Medico.id_medico))).group_by(Medico.estado).all()
    est_chart = [{"name": r[0] or "N/A", "value": r[1]} for r in est_data]

    uni_data = q.with_entities(Medico.universidad_graduacion, func.count(func.distinct(Medico.id_medico))).group_by(Medico.universidad_graduacion).all()
    uni_chart = [{"name": r[0] or "N/A", "value": r[1]} for r in uni_data]

    cen_data = q.with_entities(CentroSalud.nombre_centro, func.count(func.distinct(Medico.id_medico))).group_by(CentroSalud.nombre_centro).all()
    cen_chart = [{"name": r[0] or "N/A", "value": r[1]} for r in cen_data]

    val_data = q.with_entities(pc.c.valor_consulta_rango, func.count(func.distinct(Medico.id_medico))).group_by(pc.c.valor_consulta_rango).all()
    val_chart = [{"name": r[0] or "N/A", "value": r[1]} for r in val_data]

    pac_data = q.with_entities(pc.c.promedio_pacientes_semanal_rango, func.count(func.distinct(Medico.id_medico))).group_by(pc.c.promedio_pacientes_semanal_rango).all()
    pac_chart = [{"name": r[0] or "N/A", "value": r[1]} for r in pac_data]
    
    enc_data = q.with_entities(User.username, func.count(func.distinct(Medico.id_medico)), func.count(func.distinct(CentroSalud.id_centro)), func.count(func.distinct(EncuestaCentro.id_encuesta))).group_by(User.username).all()
    enc_ranking = [{"encuestador": r[0], "medicos": r[1], "centros": r[2], "encuestas": r[3]} for r in enc_data]

    dias_data = q.with_entities(pc.c.horarios_json).all()
    dias_count = {d: 0 for d in DIAS_ABREV}
    # Cobertura por hora: por cada franja activa (desde-hasta) se suma 1 a
    # cada hora que cubre -- muestra en qué horario hay más consultorios
    # abiertos, no solo cuántos "abren" a tal hora.
    horas_count = {h: 0 for h in range(24)}
    for (horarios_str,) in dias_data:
        if not horarios_str:
            continue
        try:
            h = json.loads(horarios_str)
        except (TypeError, ValueError):
            continue
        if not isinstance(h, dict):
            continue
        for d in DIAS_ABREV:
            info = h.get(d) or {}
            if not info.get('activo'):
                continue
            dias_count[d] += 1
            try:
                desde_h = int(str(info.get('desde', '00:00')).split(':')[0])
                hasta_h = int(str(info.get('hasta', '00:00')).split(':')[0])
            except (ValueError, IndexError):
                continue
            for hh in range(desde_h, min(hasta_h, 24)):
                if 0 <= hh < 24:
                    horas_count[hh] += 1
    dias_chart = [{"name": k, "value": v} for k, v in dias_count.items()]
    horas_chart = [{"name": f"{h:02d}:00", "value": v} for h, v in horas_count.items()]
    
    return {
        "success": True,
        "total_medicos": total_medicos,
        "total_centros": total_centros,
        "total_especialidades": total_especialidades,
        "total_estados": total_estados,
        "total_ciudades": total_ciudades,
        "total_encuestas": total_encuestas,
        "encuestas_30d": encuestas_30d,
        "medicos_con_2do_consultorio": dos_cons,
        "pct_2do_consultorio": pct_dos,
        "pct_whatsapp": pct(wa),
        "pct_email": pct(em),
        "pct_telefono": pct(tel),
        "pct_instagram": pct(ig),
        "pct_linkedin": pct(li),
        "charts": {
            "especialidades": esp_chart,
            "estados": est_chart,
            "universidades": uni_chart,
            "centros": cen_chart,
            "valor_consulta": val_chart,
            "pacientes_semana": pac_chart,
            "dias_consulta": dias_chart,
            "horas_consulta": horas_chart,
            "ranking_encuestadores": enc_ranking
        }
    }


@router.get("/api/cliente-encuestador/medicos")
def api_medicos_tabla(request: Request, q: str = "", page: int = 1, per_page: int = 25, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    check_rol_cliente_encuestador(current_user, db)
    
    base_q, pc = get_base_query(db)
    base_q = apply_filters(base_q, request, pc)

    # Deduplicar: solo la encuesta más reciente por médico
    dedup = _latest_medico_centro_subq(db)
    base_q = base_q.filter(MedicoCentroEncuesta.id_medico_centro.in_(dedup))
    
    if q.strip():
        search = f"%{q.strip()}%"
        base_q = base_q.filter(
            or_(
                Medico.id_medico_externo.ilike(search),
                Medico.apellido1.ilike(search),
                Medico.apellido2.ilike(search),
                Medico.nombre1.ilike(search),
                Medico.nombre2.ilike(search),
                Medico.especialidad.ilike(search),
                CentroSalud.nombre_centro.ilike(search)
            )
        )
        
    total = base_q.with_entities(func.count(func.distinct(Medico.id_medico))).scalar() or 0
    offset = (page - 1) * per_page
    
    rows = base_q.with_entities(
        Medico.id_medico, Medico.id_medico_externo,
        Medico.apellido1, Medico.apellido2, Medico.nombre1, Medico.nombre2,
        Medico.especialidad, Medico.sub_especialidad, Medico.universidad_graduacion,
        Medico.ciudad, Medico.estado, Medico.telefono, Medico.whatsapp, Medico.email,
        CentroSalud.nombre_centro, pc.c.valor_consulta_rango,
        pc.c.promedio_pacientes_semanal_rango, pc.c.horarios_json,
        EncuestaCentro.fecha_verificacion, User.username,
        JornadaEncuestador.latitud, JornadaEncuestador.longitud
    ).order_by(desc(EncuestaCentro.fecha_verificacion), Medico.apellido1).offset(offset).limit(per_page).all()
    
    medicos = []
    for r in rows:
        n2 = f" {r.nombre2}" if r.nombre2 else ""
        a2 = f" {r.apellido2}" if r.apellido2 else ""
        nombre_completo = f"{r.apellido1}{a2}, {r.nombre1}{n2}"
        medicos.append({
            "id_medico": r.id_medico,
            "id_medico_externo": r.id_medico_externo,
            "nombre_completo": nombre_completo,
            "especialidad": r.especialidad,
            "sub_especialidad": r.sub_especialidad,
            "universidad": r.universidad_graduacion,
            "ciudad": r.ciudad,
            "estado": r.estado,
            "telefono": r.telefono,
            "whatsapp": r.whatsapp,
            "email": r.email,
            "centro": r.nombre_centro,
            "valor_consulta_rango": r.valor_consulta_rango,
            "promedio_pacientes": r.promedio_pacientes_semanal_rango,
            "dias_consulta": _dias_activos_str(r.horarios_json),
            "fecha_verificacion": r.fecha_verificacion.isoformat() if r.fecha_verificacion else None,
            "encuestador": r.username,
            "latitud": r.latitud,
            "longitud": r.longitud
        })
        
    return {
        "success": True, "total": total, "page": page, "per_page": per_page,
        "medicos": medicos
    }


@router.get("/export")
def api_export_excel(request: Request, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    check_rol_cliente_encuestador(current_user, db)

    base_q, pc = get_base_query(db)
    base_q = apply_filters(base_q, request, pc)

    # Deduplicar: solo la encuesta más reciente por médico
    dedup = _latest_medico_centro_subq(db)
    base_q = base_q.filter(MedicoCentroEncuesta.id_medico_centro.in_(dedup))

    rows = base_q.with_entities(
        Medico.id_medico_externo, Medico.apellido1, Medico.apellido2, Medico.nombre1, Medico.nombre2,
        Medico.especialidad, Medico.sub_especialidad, Medico.universidad_graduacion,
        Medico.ciudad, Medico.estado, Medico.telefono, Medico.whatsapp, Medico.email,
        Medico.linkedin, Medico.instagram,
        CentroSalud.nombre_centro, pc.c.nombre_clinica, pc.c.piso_consultorio, pc.c.direccion_especifica,
        pc.c.valor_consulta_rango, pc.c.promedio_pacientes_semanal_rango, pc.c.horarios_json,
        EncuestaCentro.fecha_verificacion, EncuestaCentro.fuente_informacion, User.username,
    ).order_by(desc(EncuestaCentro.fecha_verificacion), Medico.apellido1).all()

    columnas = [
        "ID Externo", "Apellidos", "Nombres", "Especialidad", "Sub-especialidad", "Universidad",
        "Ciudad", "Estado", "Teléfono", "WhatsApp", "Email", "LinkedIn", "Instagram",
        "Centro de Salud", "Consultorio", "Piso/Consultorio", "Dirección",
        "Valor Consulta", "Pacientes/Semana", "Días de Consulta",
        "Fecha Verificación", "Fuente", "Encuestador",
    ]
    data = []
    for r in rows:
        data.append({
            "ID Externo": r.id_medico_externo,
            "Apellidos": f"{r.apellido1} {r.apellido2 or ''}".strip(),
            "Nombres": f"{r.nombre1} {r.nombre2 or ''}".strip(),
            "Especialidad": r.especialidad,
            "Sub-especialidad": r.sub_especialidad,
            "Universidad": r.universidad_graduacion,
            "Ciudad": r.ciudad,
            "Estado": r.estado,
            "Teléfono": r.telefono,
            "WhatsApp": r.whatsapp,
            "Email": r.email,
            "LinkedIn": r.linkedin,
            "Instagram": r.instagram,
            "Centro de Salud": r.nombre_centro,
            "Consultorio": r.nombre_clinica,
            "Piso/Consultorio": r.piso_consultorio,
            "Dirección": r.direccion_especifica,
            "Valor Consulta": r.valor_consulta_rango,
            "Pacientes/Semana": r.promedio_pacientes_semanal_rango,
            "Días de Consulta": _dias_activos_str(r.horarios_json),
            "Fecha Verificación": r.fecha_verificacion.isoformat() if r.fecha_verificacion else None,
            "Fuente": r.fuente_informacion,
            "Encuestador": r.username,
        })

    df = pd.DataFrame(data, columns=columnas) if data else pd.DataFrame(columns=columnas)

    # engine='openpyxl' -- es el que está declarado en requirements.txt.
    # xlsxwriter (usado en reporteria.py) no lo está; ese endpoint probablemente
    # ya falla en producción con ModuleNotFoundError, pero no es parte de este cambio.
    from openpyxl.styles import Font, PatternFill

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Médicos')
        worksheet = writer.sheets['Médicos']
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='21262D', end_color='21262D', fill_type='solid')
        for col_num, value in enumerate(df.columns.values, start=1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            max_len = max(df[value].astype(str).map(len).max() if not df.empty else 0, len(str(value))) + 2
            worksheet.column_dimensions[cell.column_letter].width = min(max_len, 45)
    output.seek(0)

    filename = f"IQVIA_Medicos_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        output,
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
